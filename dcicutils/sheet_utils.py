import contextlib

import chardet
import csv
import glob
import io
import json
import openpyxl
import os
import re
import subprocess
import yaml

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from tempfile import TemporaryFile, TemporaryDirectory
from typing import Any, Dict, Iterable, List, Optional, Tuple, Type, Union
from .common import AnyJsonData, Regexp, JsonSchema
from .lang_utils import conjoined_list, disjoined_list, maybe_pluralize
from .misc_utils import ignored, pad_to, JsonLinesReader, remove_suffix, to_camel_case


Header = str
Headers = List[str]
TabbedHeaders = Dict[str, Headers]
ParsedHeader = List[Union[str, int]]
ParsedHeaders = List[ParsedHeader]
TabbedParsedHeaders = Dict[str, ParsedHeaders]
SheetCellValue = Union[int, float, str]
SheetRow = List[SheetCellValue]
CsvReader = type(csv.reader(TemporaryFile()))
SheetData = List[dict]
TabbedSheetData = Dict[str, SheetData]
TabbedJsonSchemas = Dict[str, JsonSchema]
CommentRow = object()


class LoadFailure(Exception):
    """
    In general, we'd prefer to load up the spreadsheet with clumsy data that can then be validated in detail,
    but some errors are so confusing or so problematic that we need to just fail the load right away.
    """
    pass


class LoadArgumentsError(LoadFailure):
    """
    Errors of this class represent situations where we can't get started because
    there's a problem with the given arguments.
    """
    pass


class LoadTableError(LoadFailure):
    """
    Errors of this class represent situations where we can't get started because
    there's a problem with some table's syntax, for example headers that don't make sense.
    """
    pass


def unwanted_kwargs(*, context, kwargs, context_plural=False, detailed=False):
    if kwargs:
        unwanted = [f"{argname}={value!r}" if detailed else argname
                    for argname, value in kwargs.items()
                    if value is not None]
        if unwanted:
            does_not = "don't" if context_plural else "doesn't"
            raise LoadArgumentsError(f"{context} {does_not} use"
                                     f" {maybe_pluralize(unwanted, 'keyword argument')} {conjoined_list(unwanted)}.")


def prefer_number(value: SheetCellValue, kind='number'):
    """
    Given a string, if the string has number syntax, returns the number it represents. Otherwise, returns its argument.
    (It follows from this that if given an int or a float, it just returns that argument.)

    Using a kind= argument (as in kind='int' or kind='float') can better restrict what kind of number a string
    is coerced to, but it has no effect when the argument is a number, even a number of the wrong kind.

    :param value: a string, int, or float
    :param kind: one of 'number' or 'integer'
    :returns: the argument coerced to a number of the appropriate kind, if possible, or else the argument literally
    """
    if isinstance(value, str):  # the given value might be an int or float, in which case just fall through
        if not value:
            return None
        value = value
        ch0 = value[0]
        if ch0 == '+' or ch0 == '-' or ch0.isdigit():
            if kind == 'integer':
                try:
                    return int(value)
                except Exception:
                    pass
            if kind == 'number':
                try:
                    return float(value)
                except Exception:
                    pass
        # If we couldn't parse it as an int or float, fall through to returning the original value
        pass
    # NOTE WELL:
    #   In the case where we already have a number, even if it's the wrong type, we just leave it as we got it.
    #   The job of this function is not to do type enforcement or correctness checking, but rather to adjust
    #   for the fact that spreadsheets and csv files often pass string data where they mean to pass numbers.
    #   If some human has already been thinking about what to pass in a JSON or other such setting,
    #   this function is not trying to be smart enough to second-guess that.
    return value


def expand_string_escape_sequences(text: str) -> str:
    """
    Expands string escape sequences in a commonly used way.
    A backslash followed by one of the following characters is expanded as indicated:
        r (return or CR) - ASCII 13 decimal, 15 octal, 0d hex
        n (newline or linefeed or LF) - ASCII 10 decimal, 12 octal, 0a hex
        t (tab) - ASCII 9 decimal, 11 octal, 9 hex
        f (formfeed or page) - ASCII 12 decimal, 14 octal, 0c hex
        \\ (backslash) - ASCII 92 decimal, 134 octal, 5c hex
    In all other situations, the backslash is left uninterpreted.
    """
    s = io.StringIO()
    escaping = False
    for ch in text:
        if escaping:
            if ch == 'r':
                s.write('\r')
            elif ch == 't':
                s.write('\t')
            elif ch == 'n':
                s.write('\n')
            elif ch == '\\':
                s.write('\\')
            else:
                # Rather than err, just leave other sequences as-is.
                s.write(f"\\{ch}")
            escaping = False
        elif ch == '\\':
            escaping = True
        else:
            s.write(ch)
    return s.getvalue()


def open_unicode_text_input_file_respecting_byte_order_mark(filename):
    """
    Opens a file for text input, respecting a byte-order mark (BOM).
    """
    with io.open(filename, 'rb') as fp:
        leading_bytes = fp.read(4 * 8)  # 4 bytes is all we need
        bom_info = chardet.detect(leading_bytes, should_rename_legacy=True)
        detected_encoding = bom_info and bom_info.get('encoding')  # tread lightly
    use_encoding = 'utf-8' if detected_encoding == 'ascii' else detected_encoding
    return io.open(filename, 'r', encoding=use_encoding)


class AbstractTableSetManager:
    """
    The TableSetManager is the spanning class of anything that wants to be able to load a table set,
    regardless of what it wants to load it from. To do this, it must support a load method
    that takes a filename and returns the file content in the form:
        {
            "Sheet1": [
                          {...representation of row1 as some kind of dict...},
                          {...representation of row2 as some kind of dict...}
                      ],
            "Sheet2": [...],
            ...,
        }
    It also needs some implementation of the .tab_names property.
    Note that at this level of abstraction, we take no position on what form of representation is used
    for the rows, as long as it is JSON data of some kind. It might be
         {"col1": "val1", "col2": "val2", ...}
    or it might be something more structured like
         {"something": "val1", {"something_else": ["val2"]}}
    Additionally, the values stored might be altered as well. In particular, the most likely alteration
    is to turn "123" to 123 or "" to None, though the specifics of whether and how such transformations
    happen is not constrained by this class.
    """

    ALLOWED_FILE_EXTENSIONS: List[str] = []

    def __init__(self, filename: str, prefer_number: Optional[bool] = None, **kwargs):
        if prefer_number:
            # It's OK to pass prefer_number=None (meaning take the default) and prefer_number=False,
            # since that requires no action, but if a class wants to manage such preferences,
            # as happens in FlattenedTableSetManager, it will have to do it itself.
            raise ValueError(f"This class {self.__class__.__name__} does not implement prefer_number={prefer_number!r}")
        self.filename: str = filename
        unwanted_kwargs(context=self.__class__.__name__, kwargs=kwargs)

    @classmethod
    def load(cls, filename: str, **kwargs) -> TabbedSheetData:
        """
        Reads a filename and returns a dictionary that maps sheet names to rows of dictionary data.
        For more information, see documentation of AbstractTableSetManager.
        """
        raise NotImplementedError(f".load(...) is not implemented for {cls.__name__}.")  # noQA

    @property
    def tab_names(self, order: Optional[List[str]] = None) -> List[str]:
        raise NotImplementedError(f".tab_names is not implemented for {self.__class__.__name__}..")  # noQA

    def load_content(self, sheet_order: Optional[List[str]] = None) -> Any:
        raise NotImplementedError(f".load_content() is not implemented for {self.__class__.__name__}.")  # noQA


class BasicTableSetManager(AbstractTableSetManager):
    """
    A BasicTableManager provides some structure that most kinds of parsers will need.
    In particular, everything will likely need some way of storing headers and some way of storing content
    of each sheet. Even a csv file, which doesn't have multiple tabs can be seen as the degenerate case
    of this where there's only one set of headers and only one block of content.
    """

    def __init__(self, filename: str, **kwargs):
        super().__init__(filename=filename, **kwargs)
        self.headers_by_tab_name: Dict[str, Headers] = {}
        self.content_by_tab_name: Dict[str, SheetData] = {}
        self.reader_agent: Any = self._get_reader_agent()

    def tab_headers(self, tab_name: str) -> Headers:
        return self.headers_by_tab_name[tab_name]

    def tab_content(self, tab_name: str) -> List[AnyJsonData]:
        return self.content_by_tab_name[tab_name]

    @classmethod
    def _create_tab_processor_state(cls, tab_name: str) -> Any:
        """
        This method provides for the possibility that some parsers will want auxiliary state,
        (such as parsed headers or a line count or a table of temporary names for objects to cross-link
        or some other such feature) that it carries with it as it moves from line to line parsing things.
        Subclasses might therefore want to make this do something more interesting.
        """
        ignored(tab_name)  # subclasses might need this, but we don't
        return None

    def _get_reader_agent(self) -> Any:
        """This function is responsible for opening the workbook and returning a workbook object."""
        raise NotImplementedError(f"._get_reader_agent() is not implemented for {self.__class__.__name__}.")  # noQA


class FlattenedTableSetManager(BasicTableSetManager):
    """
    This is the base class for all workbook-like data sources, i.e., that may need to apply semantic processing.
    Those may be:
    * Excel workbook readers (.xlsx)
    * Comma-separated file readers (.csv)
    * Tab-separarated file readers (.tsv in most of the world, but Microsoft stupidly calls this .txt, outright
      refusing to write a .tsv file, so many people seem to compromise and call this .tsv.txt)
    There are two levels to each of these: a class that is not semantically interpreted,
    and a class that is semantically interpreted as an "item".

    This is NOT a parent class of these kinds of files, which we always take literally as if semantic processing
    were already done (in part so that they can be used to test the results of other formats):
    * Json files
    * Yaml files
    * Inserts directories
    * JsonLines files
    """

    @classmethod
    def load(cls, filename: str, **kwargs) -> AnyJsonData:
        if cls.ALLOWED_FILE_EXTENSIONS:
            if not any(filename.lower().endswith(suffix) for suffix in cls.ALLOWED_FILE_EXTENSIONS):
                raise LoadArgumentsError(f"The TableSetManager subclass {cls.__name__} expects only"
                                         f" {disjoined_list(cls.ALLOWED_FILE_EXTENSIONS)} filenames: {filename}")

        table_set_manager: FlattenedTableSetManager = cls(filename=filename, **kwargs)
        return table_set_manager.load_content()

    def __init__(self, filename: str, prefer_number: Optional[bool] = None, **kwargs):
        if prefer_number is None:  # i.e., no initial value specified
            prefer_number = True
        self.prefer_number: bool = prefer_number
        super().__init__(filename=filename, **kwargs)

    def _raw_row_generator_for_tab_name(self, tab_name: str) -> Iterable[SheetRow]:
        """
        Given a tab_name and a state (returned by _sheet_loader_state), return a generator for a set of row values.
        """
        raise NotImplementedError(f"._rows_for_tab_name(...) is not implemented for {self.__class__.__name__}.")  # noQA

    def _process_row(self, tab_name: str, state: Any, row: List[SheetCellValue]) -> AnyJsonData:
        """
        This needs to take a state and whatever represents a row and
        must return a list of objects representing column values.
        What constitutes a processed up to the class, but other than that the result must be a JSON dictionary.
        """
        raise NotImplementedError(f"._process_row(...) is not implemented for {self.__class__.__name__}.")  # noQA

    def load_content(self, sheet_order: Optional[List[str]] = None) -> AnyJsonData:
        for tab_name in self.tab_names(sheet_order):
            sheet_content = []
            state = self._create_tab_processor_state(tab_name)
            for row_data in self._raw_row_generator_for_tab_name(tab_name):
                processed_row_data: AnyJsonData = self._process_row(tab_name, state, row_data)
                if not processed_row_data:
                    break
                if processed_row_data is CommentRow:
                    continue
                sheet_content.append(processed_row_data)
            self.content_by_tab_name[tab_name.replace(" ", "")] = sheet_content
        return self.content_by_tab_name

    def parse_cell_value(self, value: SheetCellValue, override_prefer_number: Optional[bool] = None) -> AnyJsonData:
        if override_prefer_number is None:
            override_prefer_number = self.prefer_number
        return prefer_number(value) if override_prefer_number else value


class TableSetManagerRegistry:

    def __init__(self):
        self.manager_table: Dict[str, Type[AbstractTableSetManager]] = {}
        self.regexp_mappings: List[Tuple[Regexp, Type[AbstractTableSetManager]]] = []

    def register(self, regexp: Optional[str] = None):
        def _wrapped_register(class_to_register: Type[AbstractTableSetManager]):
            if regexp:
                self.regexp_mappings.append((re.compile(regexp), class_to_register))
            for ext in class_to_register.ALLOWED_FILE_EXTENSIONS:
                existing = self.manager_table.get(ext)
                if existing:
                    raise Exception(f"Tried to define {class_to_register} to extension {ext},"
                                    f" but {existing} already claimed that.")
                self.manager_table[ext] = class_to_register
            return class_to_register
        return _wrapped_register

    register1 = register

    def manager_for_filename(self, filename: str) -> Type[AbstractTableSetManager]:
        base: str = os.path.basename(filename)
        suffix_parts = base.split('.')[1:]
        if suffix_parts:
            for i in range(0, len(suffix_parts)):
                suffix = f".{'.'.join(suffix_parts[i:])}"
                found: Optional[Type[AbstractTableSetManager]] = self.manager_table.get(suffix)
                if found:
                    return found
        else:
            special_case: Optional[Type[AbstractTableSetManager]] = self.manager_for_special_filename(filename)
            if special_case:
                return special_case
        raise LoadArgumentsError(f"Unknown file type: {filename}")

    def manager_for_special_filename(self, filename: str) -> Optional[Type[AbstractTableSetManager]]:
        for pattern, manager_class in self.regexp_mappings:
            if pattern.match(filename):
                return manager_class
        return None


TABLE_SET_MANAGER_REGISTRY = TableSetManagerRegistry()


@TABLE_SET_MANAGER_REGISTRY.register()
class XlsxManager(FlattenedTableSetManager):
    """
    This implements the mechanism to get a series of rows out of the sheets in an XLSX file.
    """
    ALLOWED_FILE_EXTENSIONS = ['.xlsx']
    TERMINATE_ON_EMPTY_ROW = True
    CONVERT_VALUES_TO_STRING = True

    @classmethod
    def _all_rows(cls, sheet: Worksheet):
        row_max = sheet.max_row
        for row in range(2, row_max + 1):
            yield row

    @classmethod
    def _all_cols(cls, sheet: Worksheet):
        col_max = sheet.max_column
        for col in range(1, col_max + 1):
            yield col

    def tab_names(self, order: Optional[List[str]] = None) -> List[str]:
        def ordered_sheet_names(sheet_names: List[str]) -> List[str]:
            if not order:
                return sheet_names
            ordered_sheet_names = []
            for item in order:
                for sheet_name in sheet_names:
                    if to_camel_case(item.replace(" ", "")) == to_camel_case(sheet_name.replace(" ", "")):
                        ordered_sheet_names.append(sheet_name)
            for sheet_name in sheet_names:
                if sheet_name not in ordered_sheet_names:
                    ordered_sheet_names.append(sheet_name)
            return ordered_sheet_names
        return ordered_sheet_names(self.reader_agent.sheetnames)

    def _get_reader_agent(self) -> Workbook:
        return openpyxl.load_workbook(self.filename)

    def _raw_row_generator_for_tab_name(self, tab_name: str) -> Iterable[SheetRow]:
        sheet = self.reader_agent[tab_name]
        return (self._get_raw_row_content_tuple(sheet, row)
                for row in self._all_rows(sheet))

    def _get_raw_row_content_tuple(self, sheet: Worksheet, row: int) -> SheetRow:
        # return [str(sheet.cell(row=row, column=col).value) for col in self._all_cols(sheet)]
        results = []
        for col in self._all_cols(sheet):
            value = sheet.cell(row=row, column=col).value
#           if value is not None and XlsxManager.CONVERT_VALUES_TO_STRING:
#               value = str(value).strip()
            if XlsxManager.CONVERT_VALUES_TO_STRING:
                value = str(value).strip() if value is not None else ""
            results.append(value)
        return results

    def _create_tab_processor_state(self, tab_name: str) -> Headers:
        sheet = self.reader_agent[tab_name]
        """
        headers: Headers = [str(sheet.cell(row=1, column=col).value)
                            for col in self._all_cols(sheet)]
        """
        headers = []
        for col in self._all_cols(sheet):
            cell = sheet.cell(row=1, column=col).value
            if cell is not None and XlsxManager.CONVERT_VALUES_TO_STRING:
                cell = str(cell).strip()
            headers.append(cell)
        self.headers_by_tab_name[sheet.title] = XlsxManager.remove_trailing_none_values(headers)
        return self.headers_by_tab_name[sheet.title]

    def _process_row(self, tab_name: str, headers: Headers, row_data: SheetRow) -> AnyJsonData:
        ignored(tab_name)
        if XlsxManager.is_terminating_row(row_data):
            return None
        if XlsxManager.is_comment_row(row_data):
            return CommentRow
        if len(headers) < len(row_data):
            row_data = row_data[:len(headers)]
        override_prefer_number = False if XlsxManager.CONVERT_VALUES_TO_STRING else None
        return {headers[i]: self.parse_cell_value(row_datum, override_prefer_number=override_prefer_number)
                for i, row_datum in enumerate(row_data)}

    @staticmethod
    def is_terminating_row(row: List[Optional[Any]]) -> bool:
        # TODO: This is may change; currently an all blank row signals the end of input.
        return all(cell is None or cell == "" for cell in row) and XlsxManager.TERMINATE_ON_EMPTY_ROW

    @staticmethod
    def is_comment_row(row: Tuple[Optional[Any]]) -> bool:
        # TODO: This will probably change; currently a row starting only with #, *, or ^ signals a comment.
        for cell in row:
            if cell is None:
                continue
            if (cell := str(cell)).startswith("#") or cell.startswith("*") or cell.startswith("^"):
                return True
        return False

    @staticmethod
    def remove_trailing_none_values(values: List[Any]) -> List[Any]:
        for index in range(len(values) - 1, -1, -1):
            if values[index] is not None:
                break
        else:
            return []
        return values[:index + 1]


def infer_tab_name_from_filename(filename):
    return os.path.basename(filename).split('.')[0]


class SingleTableMixin(AbstractTableSetManager):

    def __init__(self, filename: str, tab_name: Optional[str] = None, **kwargs):
        self._tab_name = tab_name or infer_tab_name_from_filename(filename)
        super().__init__(filename=filename, **kwargs)

    def tab_names(self, order: Optional[List[str]] = None) -> List[str]:
        return [self._tab_name]


class InsertsManager(BasicTableSetManager):

    ALLOWED_FILE_EXTENSIONS = []

    def _parse_inserts_data(self, filename: str) -> AnyJsonData:
        raise NotImplementedError(f"._parse_inserts_dataa(...) is not implemented for {self.__class__.__name__}.")  # noQA

    def _load_inserts_data(self, filename: str) -> TabbedSheetData:
        raw_data: AnyJsonData = self._parse_inserts_data(filename)
        tabbed_inserts: AnyJsonData = self._wrap_inserts_data(filename, raw_data)
        if (not isinstance(tabbed_inserts, dict)
                or not all(isinstance(tab_name, str) for tab_name in tabbed_inserts.keys())
                or not all(isinstance(data, list) and all(isinstance(datum, dict) for datum in data)
                           for data in tabbed_inserts.values())):
            raise ValueError(f"Data in {filename} is not of type TabbedSheetData (Dict[str, List[dict]]).")
        tabbed_inserts: TabbedSheetData  # we've just checked that
        return tabbed_inserts

    @classmethod
    def _wrap_inserts_data(cls, filename: str, data: AnyJsonData) -> AnyJsonData:
        ignored(filename)
        return data

    def tab_names(self, order: Optional[List[str]] = None) -> List[str]:
        return list(self.content_by_tab_name.keys())

    def _get_reader_agent(self) -> Any:
        return self

    @classmethod
    def extract_tabbed_headers(cls, data: TabbedSheetData) -> TabbedHeaders:
        result: TabbedHeaders = {}
        for tab, rows in data.items():
            if rows:
                # Data is homogeneous, so whatever the headers for the first row should be the same for all
                headers: List[str] = list(rows[0].keys())
            else:
                # If there's no data in the tab, there are also no headers
                headers: List[str] = []
            result[tab] = headers
        return result

    def load_content(self, sheet_order: Optional[List[str]] = None) -> Dict[str, AnyJsonData]:
        data = self._load_inserts_data(self.filename)
        self.content_by_tab_name = data
        self.headers_by_tab_name = self.extract_tabbed_headers(data)
        return self.content_by_tab_name


class SimpleInsertsMixin(SingleTableMixin):

    def _wrap_inserts_data(self, filename: str, data: AnyJsonData) -> TabbedSheetData:
        if (not isinstance(data, list)
                or not all(isinstance(datum, dict) for datum in data)):
            raise ValueError(f"Data in {filename} is not of type SheetData (List[dict]).")
        return {self._tab_name: data}


class JsonInsertsMixin:

    @classmethod
    def _parse_inserts_data(cls, filename: str) -> AnyJsonData:
        return json.load(open_unicode_text_input_file_respecting_byte_order_mark(filename))


@TABLE_SET_MANAGER_REGISTRY.register()
class TabbedJsonInsertsManager(JsonInsertsMixin, InsertsManager):

    ALLOWED_FILE_EXTENSIONS = [".tabs.json"]  # If you want them all in one family, use this extension


@TABLE_SET_MANAGER_REGISTRY.register()
class SimpleJsonInsertsManager(SimpleInsertsMixin, JsonInsertsMixin, InsertsManager):

    ALLOWED_FILE_EXTENSIONS = [".json"]


class YamlInsertsMixin:

    def _parse_inserts_data(self, filename) -> AnyJsonData:
        return yaml.safe_load(open_unicode_text_input_file_respecting_byte_order_mark(filename))


@TABLE_SET_MANAGER_REGISTRY.register()
class TabbedYamlInsertsManager(YamlInsertsMixin, InsertsManager):

    ALLOWED_FILE_EXTENSIONS = [".tabs.yaml"]

    def _parse_inserts_data(self, filename) -> AnyJsonData:
        return yaml.safe_load(open_unicode_text_input_file_respecting_byte_order_mark(filename))


@TABLE_SET_MANAGER_REGISTRY.register()
class SimpleYamlInsertsManager(SimpleInsertsMixin, YamlInsertsMixin, InsertsManager):

    ALLOWED_FILE_EXTENSIONS = [".yaml"]


@TABLE_SET_MANAGER_REGISTRY.register()
class SimpleJsonLinesInsertsManager(SimpleInsertsMixin, InsertsManager):

    ALLOWED_FILE_EXTENSIONS = [".jsonl"]

    def _parse_inserts_data(self, filename: str) -> AnyJsonData:
        return [line for line in JsonLinesReader(open_unicode_text_input_file_respecting_byte_order_mark(filename))]


@TABLE_SET_MANAGER_REGISTRY.register(regexp="^(.*/)?(|[^/]*[-_])inserts/?$")
class InsertsDirectoryManager(InsertsManager):

    ALLOWED_FILE_EXTENSIONS = []

    def _parse_inserts_data(self, filename: str) -> AnyJsonData:
        if not os.path.isdir(filename):
            raise LoadArgumentsError(f"{filename} is not the name of an inserts directory.")
        tab_files = glob.glob(os.path.join(filename, "*.json"))
        data = {}
        for tab_file in tab_files:
            tab_content = json.load(open_unicode_text_input_file_respecting_byte_order_mark(tab_file))
            # Here we don't use os.path.splitext because we want to split on the first dot.
            # e.g., for foo.bar.baz, return just foo
            #       this allows names like ExperimentSet.tab.json that might need to use multi-dot suffixes
            #       for things unrelated to the tab name.
            tab_name = os.path.basename(tab_file).split('.')[0]
            data[tab_name] = tab_content
        return data


@TABLE_SET_MANAGER_REGISTRY.register()
class CsvManager(SingleTableMixin, FlattenedTableSetManager):
    """
    This implements the mechanism to get a series of rows out of the sheet in a csv file,
    returning a result that still looks like there could have been multiple tabs.
    """

    ALLOWED_FILE_EXTENSIONS = ['.csv']

    def __init__(self, filename: str, escaping: Optional[bool] = None, **kwargs):
        super().__init__(filename=filename, **kwargs)
        self.escaping: bool = escaping or False

    def _get_reader_agent(self) -> CsvReader:
        return self._get_reader_agent_for_filename(self.filename)

    @classmethod
    def _get_reader_agent_for_filename(cls, filename) -> CsvReader:
        return csv.reader(open_unicode_text_input_file_respecting_byte_order_mark(filename))

    PAD_TRAILING_TABS = True

    def _raw_row_generator_for_tab_name(self, tab_name: str) -> Iterable[SheetRow]:
        headers = self.tab_headers(tab_name)
        n_headers = len(headers)
        for row_data in self.reader_agent:
            if self.PAD_TRAILING_TABS:
                row_data = pad_to(n_headers, row_data, padding='')
            yield row_data

    def _create_tab_processor_state(self, tab_name: str) -> Headers:
        headers: Optional[Headers] = self.headers_by_tab_name.get(tab_name)
        if headers is None:
            self.headers_by_tab_name[tab_name] = headers = self.reader_agent.__next__()
        return headers

    @classmethod
    def _escape_cell_text(cls, cell_text):
        if '\\' in cell_text:
            return expand_string_escape_sequences(cell_text)
        else:
            return cell_text

    def _process_row(self, tab_name: str, headers: Headers, row_data: SheetRow) -> AnyJsonData:
        ignored(tab_name)
        if self.escaping:
            return {headers[i]: self.parse_cell_value(self._escape_cell_text(cell_text), override_prefer_number=False)
                    for i, cell_text in enumerate(row_data)}
        else:
            return {headers[i]: self.parse_cell_value(cell_text, override_prefer_number=False)
                    for i, cell_text in enumerate(row_data)}


@TABLE_SET_MANAGER_REGISTRY.register()
class TsvManager(CsvManager):
    """
    TSV files are just CSV files with tabs instead of commas as separators.
    (We do not presently handle any escaping of strange characters. May need to add handling for backslash escaping.)
    """
    ALLOWED_FILE_EXTENSIONS = ['.tsv', '.tsv.txt']

    @classmethod
    def _get_reader_agent_for_filename(cls, filename) -> CsvReader:
        return csv.reader(open_unicode_text_input_file_respecting_byte_order_mark(filename), delimiter='\t')


def _do_shell_command(command, cwd=None):
    # This might need to be more elaborate, but hopefully it will do for now. -kmp 11-Sep-2023
    subprocess.check_output(command, cwd=cwd)


@contextlib.contextmanager
def maybe_unpack(filename):  # Maybe move to another module
    """
    If necessary, unpack a file that is zipped and/or tarred, yielding the name of the file (unpacked or not).
    """
    unpackables = ['.tar.gz', '.tar', '.tgz', '.gz', '.zip']
    ext = None
    for unpackable in unpackables:
        if filename.endswith(unpackable):
            ext = unpackable
            break
    if not ext:
        yield filename
        return
    if not os.path.exists(filename):
        # We don't bother to raise this error if we're not planning to do any unpacking.
        # The caller can decide if/when such errors are needed in that case.
        # But if we are going to have to move bits around, they'll need to actually be there.
        # -kmp 12-Sep-2023
        raise ValueError(f"The file {filename!r} does not exist.")
    target_base_part = remove_suffix(ext, os.path.basename(filename), required=True)
    target_ext = '.tar.gz' if ext == '.tgz' else ext
    with TemporaryDirectory() as temp_dir:
        temp_base = os.path.join(temp_dir, target_base_part)
        temp_filename = temp_base + target_ext
        _do_shell_command(['cp', filename, temp_filename])
        if temp_filename.endswith('.gz'):
            _do_shell_command(['gunzip', temp_filename], cwd=temp_dir)
            temp_filename = remove_suffix('.gz', temp_filename)
        elif temp_filename.endswith(".zip"):
            _do_shell_command(['unzip', temp_filename], cwd=temp_dir)
            temp_filename = remove_suffix('.zip', temp_filename)
        if temp_filename.endswith(".tar"):
            _do_shell_command(['tar', '-xf', temp_filename], cwd=temp_dir)
            tar_file = temp_filename
            temp_filename = remove_suffix(".tar", temp_filename, required=True)
            if not os.path.isdir(temp_filename):
                raise Exception(f"{tar_file} didn't unpack to a dir: {temp_filename}")
        # print(f"Unpacked {filename} to {temp_filename}")
        yield temp_filename


class TableSetManager(AbstractTableSetManager):
    """
    This class will open a .xlsx or .csv file and load its content in our standard format.
    (See more detailed description in AbstractTableManager.)
    """

    COMPRESSION_EXTENSIONS = ['.gz', '.tgz', '.tar.gz']

    @classmethod
    def create_implementation_manager(cls, filename: str, **kwargs) -> AbstractTableSetManager:

        reader_agent_class = TABLE_SET_MANAGER_REGISTRY.manager_for_filename(filename)
        # This is a bad forward reference in current refactor, but also may be testing for something we don't need
        # to worry about anymore. -kmp 11-Sep-2023
        #
        # if issubclass(reader_agent_class, AbstractItemManager):
        #     raise ValueError(f"TableSetManager unexpectedly found reader agent class {reader_agent_class}.")
        reader_agent = reader_agent_class(filename=filename, **kwargs)
        return reader_agent

    @classmethod
    def load(cls, filename: str, tab_name: Optional[str] = None, escaping: Optional[bool] = None,
             **kwargs) -> TabbedSheetData:
        """
        Given a filename and various options
        """
        annotated_content = cls.load_annotated(filename=filename, tab_name=tab_name, escaping=escaping, **kwargs)
        content: TabbedSheetData = annotated_content['content']
        return content

    @classmethod
    def load_annotated(cls, filename: str, tab_name: Optional[str] = None, escaping: Optional[bool] = None,
                       retain_empty_properties: bool = False, sheet_order: Optional[List[str]] = None,
                       **kwargs) -> Dict:
        """
        Given a filename and various options
        """
        orig_filename = filename
        with maybe_unpack(filename) as filename:
            manager = cls.create_implementation_manager(filename=filename, tab_name=tab_name, escaping=escaping,
                                                        **kwargs)
            content: TabbedSheetData = manager.load_content(sheet_order)
            return {
                'filename': filename,
                'content': content,
                'tab_name': tab_name,
                'escaping': escaping,
                'singleton': isinstance(manager, SingleTableMixin),
                'flattened': isinstance(manager, FlattenedTableSetManager),
                'packed': orig_filename != filename,  # tar or zip file that had to be unpacked somehow
            }


load_table_set = TableSetManager.load
load_table_annotated = TableSetManager.load_annotated
