import abc
import csv
import openpyxl
from typing import Any, Generator, Iterator, List, Optional, Type, Tuple, Union
from dcicutils.misc_utils import create_dict, right_trim

# Forward type references for type hints.
Excel = Type["Excel"]

# Cell values(s) indicating property deletion.
_CELL_DELETION_VALUES = ["*delete*"]


# Special cell deletion sentinel value (note make sure on deepcopy it remains the same).
class _CellDeletionSentinal(str):
    def __new__(cls):
        return super(_CellDeletionSentinal, cls).__new__(cls, _CELL_DELETION_VALUES[0])
    def __deepcopy__(self, memo):  # noqa
        return self


class RowReader(abc.ABC):

    CELL_DELETION_SENTINEL = _CellDeletionSentinal()

    def __init__(self):
        self.header = None
        self.row_number = 0
        self._warning_empty_headers = False
        self._warning_extra_values = []  # Line numbers.
        self.open()

    def __iter__(self) -> Iterator:
        for row in self.rows:
            self.row_number += 1
            if self.is_comment_row(row):
                continue
            if self.is_terminating_row(row):
                break
            if len(self.header) < len(row):  # Row values beyond what there are headers for are ignored.
                self._warning_extra_values.append(self.row_number)
            yield {column: self.cell_value(value) for column, value in zip(self.header, row)}

    def _define_header(self, header: List[Optional[Any]]) -> None:
        self.header = []
        for index, column in enumerate(header or []):
            if not (column := str(column).strip() if column is not None else ""):
                self._warning_empty_headers = True
                break  # Empty header column signals end of header.
            self.header.append(column)

    def rows(self) -> Generator[Union[List[Optional[Any]], Tuple[Optional[Any], ...]], None, None]:
        yield

    def is_comment_row(self, row: Union[List[Optional[Any]], Tuple[Optional[Any]]]) -> bool:
        return False

    def is_terminating_row(self, row: Union[List[Optional[Any]], Tuple[Optional[Any]]]) -> bool:
        return False

    def cell_value(self, value: Optional[Any]) -> str:
        if value is None:
            return ""
        elif (value := str(value).strip()) in _CELL_DELETION_VALUES:
            return RowReader.CELL_DELETION_SENTINEL
        else:
            return value

    def open(self) -> None:
        pass

    @property
    def file(self) -> Optional[str]:
        return self._file if hasattr(self, "_file") else None

    @property
    def warnings(self) -> List[str]:
        warnings = []
        if self._warning_empty_headers:
            warnings.append({"src": create_dict(file=self.file),
                             "warning": "Empty header column encountered; ignoring it and all subsequent columns."})
        if self._warning_extra_values:
            for row_number in self._warning_extra_values:
                warnings.append({"src": create_dict(file=self.file, row=row_number),
                                 "warning": f"Extra row column values."})
        return warnings


class ListReader(RowReader):

    def __init__(self, rows: List[List[Optional[Any]]]) -> None:
        self._rows = rows
        super().__init__()

    @property
    def rows(self) -> Generator[List[Optional[Any]], None, None]:
        for row in self._rows[1:]:
            yield row

    def open(self) -> None:
        if not self.header:
            self._define_header(self._rows[0] if self._rows else [])


class CsvReader(RowReader):

    def __init__(self, file: str) -> None:
        self._file = file
        self._file_handle = None
        self._rows = None
        super().__init__()

    @property
    def rows(self) -> Generator[List[Optional[Any]], None, None]:
        for row in self._rows:
            yield right_trim(row)

    def open(self) -> None:
        if self._file_handle is None:
            self._file_handle = open(self._file)
            self._rows = csv.reader(self._file_handle, delimiter="\t" if self._file.endswith(".tsv") else ",")
            self._define_header(right_trim(next(self._rows, [])))

    def __del__(self) -> None:
        if (file_handle := self._file_handle) is not None:
            self._file_handle = None
            file_handle.close()


class ExcelSheetReader(RowReader):

    def __init__(self, excel: Excel, sheet_name: str, workbook: openpyxl.workbook.workbook.Workbook) -> None:
        self.sheet_name = sheet_name or "Sheet1"
        self._workbook = workbook
        self._file = excel._file
        self._rows = None
        super().__init__()

    @property
    def rows(self) -> Generator[Tuple[Optional[Any], ...], None, None]:
        for row in self._rows(min_row=2, values_only=True):
            yield right_trim(row)

    def is_terminating_row(self, row: Tuple[Optional[Any]]) -> bool:
        return all(cell is None for cell in row)  # Empty row signals end of data.

    def open(self) -> None:
        if not self._rows:
            self._rows = self._workbook[self.sheet_name].iter_rows
            self._define_header(right_trim(next(self._rows(min_row=1, max_row=1, values_only=True), [])))


class Excel:

    def __init__(self, file: str, reader_class: Optional[Type] = None, include_hidden_sheets: bool = False) -> None:
        self._file = file
        self._workbook = None
        self._include_hidden_sheets = include_hidden_sheets
        self.sheet_names = None
        if isinstance(reader_class, Type) and issubclass(reader_class, ExcelSheetReader):
            self._reader_class = reader_class
        else:
            self._reader_class = ExcelSheetReader
        self.open()

    def sheet_reader(self, sheet_name: str) -> ExcelSheetReader:
        return self._reader_class(self, sheet_name=sheet_name, workbook=self._workbook)

    def open(self) -> None:
        if self._workbook is None:
            import warnings
            with warnings.catch_warnings():
                # Without this warning suppression thing, for some spreadsheets we get this stdout warning:
                # UserWarning: data validation extension is not supported and will be removed
                warnings.filterwarnings("ignore", category=UserWarning)
                self._workbook = openpyxl.load_workbook(self._file, data_only=True)
            self.sheet_names = [sheet_name for sheet_name in self._workbook.sheetnames
                                if not self.is_hidden_sheet(self._workbook[sheet_name])]

    def is_hidden_sheet(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
        if not self._include_hidden_sheets:
            if sheet.sheet_state == "hidden":
                return True
            if sheet.title.startswith("(") and sheet.title.endswith(")"):
                return True
        return False

    def __del__(self) -> None:
        if (workbook := self._workbook) is not None:
            self._workbook = None
            workbook.close()
